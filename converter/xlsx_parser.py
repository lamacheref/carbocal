from __future__ import annotations

import re
from datetime import date, datetime, time

from openpyxl import load_workbook

from .model import RawRow

FIELD_ALIASES = {
    "sujet": "subject",
    "subject": "subject",
    "titre": "subject",
    "date": "date",
    "jour": "date",
    "heuredebut": "start",
    "debut": "start",
    "start": "start",
    "heurefin": "end",
    "fin": "end",
    "end": "end",
    "lieu": "location",
    "location": "location",
    "invites": "attendees",
    "attendees": "attendees",
    "attendee": "attendees",
}

_WEEKDAY_FR = r"(?:lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)"
_DATE_RE = re.compile(
    rf"^\s*(\d{{1,2}})/(\d{{1,2}})/(\d{{4}})(?:\s*[-–]\s*{_WEEKDAY_FR})?\s*$",
    re.IGNORECASE,
)


def _normalize_header(name: str) -> str:
    raw = name.strip().lower()
    raw = re.sub(r"[_\-\s]+", "", raw)
    raw = raw.replace("é", "e").replace("è", "e").replace("ê", "e").replace("à", "a")
    raw = raw.replace("ô", "o").replace("î", "i").replace("û", "u").replace("ç", "c")
    return FIELD_ALIASES.get(raw, raw)


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    m = _DATE_RE.match(str(value))
    if not m:
        raise ValueError(f"date illisible : {value!r}")
    day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    return date(year, month, day)


def _parse_time(value) -> time:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        total_minutes = int(round(value * 24 * 60)) % (24 * 60)
        return time(total_minutes // 60, total_minutes % 60)
    raw = str(value).strip()
    parts = re.split(r"[:hH]", raw)
    if len(parts) < 2:
        raise ValueError(f"heure illisible : {value!r}")
    return time(int(parts[0]), int(parts[1]))


def _parse_attendees(value) -> list[str]:
    if value is None:
        return []
    emails = []
    for part in str(value).split(";"):
        email = part.strip()
        if email:
            emails.append(email)
    return emails


def parse_workbook(path: str, sheet_name: str | None = None) -> tuple[list[RawRow], list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    errors: list[str] = []

    rows_iter = ws.iter_rows(values_only=True)
    header = next((r for r in rows_iter if any(v is not None for v in r)), None)
    if header is None:
        return [], ["feuille vide"]

    fields = {}
    for idx, value in enumerate(header):
        if value is None:
            continue
        name = _normalize_header(str(value))
        if name in {
            "subject",
            "date",
            "start",
            "end",
            "location",
            "attendees",
        }:
            fields[name] = idx

    missing = [name for name in ("subject", "date", "start", "end") if name not in fields]
    if missing:
        return [], [f"colonnes obligatoires absentes : {', '.join(missing)}"]

    rows: list[RawRow] = []
    line = 1
    for values in rows_iter:
        line += 1
        subject = values[fields["subject"]]
        if subject is None or not str(subject).strip():
            if any(v is not None and str(v).strip() for v in values):
                errors.append(f"ligne {line} : sujet vide, ignorée")
            continue
        try:
            d = _parse_date(values[fields["date"]])
            start_raw = values[fields["start"]]
            end_raw = values[fields["end"]]
            start = _parse_time(start_raw) if start_raw is not None else None
            end = _parse_time(end_raw) if end_raw is not None else None
        except ValueError as exc:
            errors.append(f"ligne {line} : {exc}")
            continue
        if (start is None) != (end is None):
            errors.append(f"ligne {line} : heure début ou fin manquante, ignorée")
            continue
        if start is not None and end is not None and end <= start:
            errors.append(f"ligne {line} : fin ({end}) <= début ({start}), ignorée")
            continue
        location = ""
        if "location" in fields and values[fields["location"]]:
            location = str(values[fields["location"]]).strip()
        attendees = []
        if "attendees" in fields:
            attendees = _parse_attendees(values[fields["attendees"]])
        rows.append(RawRow(line, str(subject).strip(), d, start, end, location, attendees))

    return rows, errors